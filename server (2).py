

import openpyxl
from openpyxl import load_workbook

import socket
import sys
from decimal import *
import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook
from sklearn.metrics import accuracy_score
from sklearn.ensemble import RandomForestClassifier
from sklearn.datasets import make_classification
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split
from sklearn.naive_bayes import GaussianNB,MultinomialNB,BernoulliNB
from sklearn.neighbors import KNeighborsClassifier
from skmultilearn.problem_transform import LabelPowerset
from sklearn.neural_network import MLPClassifier
import random
getcontext().prec=25
HOST = '192.168.43.107'  # Symbolic name, meaning all available interfaces
PORT = 8883 #Arbitrary non-privileged port

s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
print('Socket created')

#Bind socket to local host and port
try:
     s.bind((HOST, PORT))
except socket.error as msg:
     print('Bind failed. Error Code : ' + str(msg[0]) + ' Message ' + msg[1])
     sys.exit()

print('Socket bind complete')

# Start listening on socket
s.listen(10)
print('Socket now listening')

conn, addr = s.accept()
print('Connected with ' + addr[0] + ':' + str(addr[1]))
msg = conn.recv(1024)

msg=msg.decode()

features=msg.split("#")
print(features)
variance = (features[0])

idm =Decimal(features[1])
mean = Decimal(features[2])
stdx = Decimal(features[3])
stdy = Decimal(features[4])
correlation = Decimal(features[5])
contrast = Decimal(features[6])
energy = Decimal(features[7])
entropy = Decimal(features[8])
homogenity = Decimal(features[9])
shade = Decimal(features[10])
prominence = Decimal(features[11])
inertia=Decimal(features[12])
red = Decimal(features[13])
green = Decimal(features[14])
blue = Decimal(features[15])
gbyr = Decimal(features[16])
gbyb = Decimal(features[17])

#
#
# wb = load_workbook('sample.xlsx',keep_vba=True,data_only=True)
# sheet =wb.active
# rows = ((red, green, blue,gbyr,gbyb),)
#
#
# for row in rows:
#      sheet.append(row)
#
# wb.save("sample.xlsx")
#

df1=pd.read_csv("/home/mustafa/Downloads/new_data.csv")

factor = pd.factorize(df1['class'])
df1['class'] = factor[0]
definitions = factor[1]

X = df1.iloc[:,0:5].values
y = df1.iloc[:,5].values
#
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size = 0.2,random_state = 22)
scaler = StandardScaler()
X_train = scaler.fit_transform(X_train)
X_test = scaler.transform(X_test)

neigh = KNeighborsClassifier(n_neighbors=5)
neigh.fit(X_train, y_train)
#
#
# #
clf = (GaussianNB())
clf.fit(X_train, y_train)

clef = MLPClassifier(solver='lbfgs', alpha=1e-5,hidden_layer_sizes=(9, 9), random_state=1)
clef.fit(X, y)


classifier = RandomForestClassifier(n_estimators=3,random_state=45)
classifier.fit(X_train, y_train)
y_pred = clef.predict([[red,green,blue,gbyr,gbyb]])
#y_pred = classifier.predict([[idm,mean,stdx,stdy,correlation,contrast,energy,entropy,homogenity,shade,prominence,inertia]])
#y_pred = neigh.predict([[idm,mean,stdx,stdy,correlation,contrast,energy,entropy,homogenity,shade,prominence,inertia]])
#Reverse factorize (converting y_pred from 0s,1s and 2s to Iris-setosa, Iris-versicolor and Iris-virginica
reversefactor = dict(zip(range(3),definitions))
y_test = np.vectorize(reversefactor.get)(y_test)
y_pred = np.vectorize(reversefactor.get)(y_pred)
# # Making the Confusion Matrix
# #print(pd.crosstab(y_test, y_pred))
print(y_pred)
# #print(accuracy_score(y_test, y_pred,normalize=True))
# answer = y_pred
conn.sendall(y_pred)
conn.close()
s.close()
